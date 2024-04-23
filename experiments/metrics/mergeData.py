import pandas as pd
import openpyxl
from openpyxl.styles import Alignment


# unbounded
def mf(args, output_file,exec_time_flag):
    headers=[]
    # unbounded
    workbook = openpyxl.load_workbook(args[0])
    sheet = workbook.active
    sheet.insert_rows(1)
    if exec_time_flag:
        sheet.merge_cells("A1:C1")
    else:
        sheet.merge_cells("A1:D1")
    sheet["A1"] = "Unbounded"
    cell = sheet["A1"]
    cell.alignment = Alignment(horizontal="center")
    workbook.save(args[0])
    # bound 0
    workbook0 = openpyxl.load_workbook(args[1])
    sheet0 = workbook0.active
    sheet0.insert_rows(1)
    if exec_time_flag:
        sheet0.merge_cells("A1:C1")
    else:
        sheet0.merge_cells("A1:D1")
    sheet0["A1"] = "Bound 0"
    cell = sheet0["A1"]
    cell.alignment = Alignment(horizontal="center")
    workbook0.save(args[1])
    # bound 1
    workbook1 = openpyxl.load_workbook(args[2])
    sheet1 = workbook1.active
    sheet1.insert_rows(1)
    if exec_time_flag:
        sheet1.merge_cells("A1:C1")
    else:
        sheet1.merge_cells("A1:D1")
    sheet1["A1"] = "Bound 1"
    cell = sheet1["A1"]
    cell.alignment = Alignment(horizontal="center")
    workbook1.save(args[2])
    # bound 2
    workbook2 = openpyxl.load_workbook(args[3])
    sheet2 = workbook2.active
    sheet2.insert_rows(1)
    if exec_time_flag:
        sheet2.merge_cells("A1:C1")
    else:
        sheet2.merge_cells("A1:D1")
    sheet2["A1"] = "Bound 2"
    cell = sheet2["A1"]
    cell.alignment = Alignment(horizontal="center")
    workbook2.save(args[3])
    # bound 3
    workbook3 = openpyxl.load_workbook(args[4])
    sheet3 = workbook3.active
    sheet3.insert_rows(1)
    if exec_time_flag:
        sheet3.merge_cells("A1:C1")
    else:
        sheet3.merge_cells("A1:D1")
    sheet3["A1"] = "Bound 3"
    cell = sheet3["A1"]
    cell.alignment = Alignment(horizontal="center")
    workbook3.save(args[4])
    # bound 4
    workbook4 = openpyxl.load_workbook(args[5])
    sheet4 = workbook4.active
    sheet4.insert_rows(1)
    if exec_time_flag:
        sheet4.merge_cells("A1:C1")
    else:
        sheet4.merge_cells("A1:D1")
    sheet4["A1"] = "Bound 4"
    cell = sheet4["A1"]
    cell.alignment = Alignment(horizontal="center")
    workbook4.save(args[5])
    # bound 5
    workbook5 = openpyxl.load_workbook(args[6])
    sheet5 = workbook5.active
    sheet5.insert_rows(1)
    if exec_time_flag:
        sheet5.merge_cells("A1:C1")
    else:
        sheet5.merge_cells("A1:D1")
    sheet5["A1"] = "Bound 5"
    cell = sheet5["A1"]
    cell.alignment = Alignment(horizontal="center")
    workbook5.save(args[6])

    df_unb = pd.read_excel(args[0], header=None)
    df_0 = pd.read_excel(args[1], header=None)
    df_1 = pd.read_excel(args[2], header=None)
    df_2 = pd.read_excel(args[3], header=None)
    df_3 = pd.read_excel(args[4], header=None)
    df_4 = pd.read_excel(args[5], header=None)
    df_5 = pd.read_excel(args[6], header=None)

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
        merged_df = pd.concat([df_unb, df_0, df_1, df_2, df_3, df_4, df_5], axis=1, ignore_index=True)
        merged_df.to_excel(writer, sheet_name="All", startrow=0, startcol=0, header=False, index=False)

        
        merged_df = pd.concat(
            [
                df_unb.iloc[:, :4],
                df_0.iloc[:, :4],
                df_1.iloc[:, :4],
                df_2.iloc[:, :4],
                df_3.iloc[:, :4],
                df_4.iloc[:, :4],
                df_5.iloc[:, :4],
            ],
            axis=1,
        )
        merged_df.to_excel(writer, sheet_name="CallSite", startrow=0, startcol=0, header=False, index=False)
        sheetcallsite = writer.sheets["CallSite"]
        # Define the headers and merge cells as needed for 'Sheet2'
        headers = [
            "Unbounded",
            "Bound 0",
            "Bound 1",
            "Bound 2",
            "Bound 3",
            "Bound 4",
            "Bound 5",
        ]
        if exec_time_flag:
            col_ranges = ["A1:D1", "E1:H1", "I1:L1", "M1:P1", "Q1:T1", "U1:X1", "Y1:AB1"]
        else:
            col_ranges = ["A1:C1", "D1:F1", "G1:I1", "J1:L1", "M1:O1", "P1:R1", "S1:U1"]

        for i, header in enumerate(headers):
            sheetcallsite.merge_cells(col_ranges[i])
            sheetcallsite[col_ranges[i].split(":")[0]] = header
            cell = sheetcallsite[col_ranges[i].split(":")[0]]
            cell.alignment = Alignment(horizontal="center")

        merged_df = pd.concat(
    [
        df.iloc[:, :2].join(df.iloc[:, -2:], rsuffix="_suffix")
        for df in [df_unb, df_0, df_1, df_2, df_3, df_4, df_5]
    ],
    axis=1,
    )

        print("Final output written to", output_file)
